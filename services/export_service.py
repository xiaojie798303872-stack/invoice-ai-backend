# -*- coding: utf-8 -*-
"""
导出服务模块
支持将发票数据导出为Excel和PDF格式
"""

import os
import io
import logging
from typing import List, Optional
from datetime import datetime

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models import Invoice
from config import settings

logger = logging.getLogger(__name__)


class ExportService:
    """发票导出服务类"""

    async def export_to_excel(
        self,
        db: AsyncSession,
        invoices: List[Invoice],
        output_path: Optional[str] = None,
    ) -> str:
        """
        将发票数据导出为Excel文件

        Args:
            db: 数据库会话
            invoices: 发票列表
            output_path: 输出文件路径，为None时自动生成

        Returns:
            str: 导出文件的路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("openpyxl未安装，请执行: pip install openpyxl")

        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(
                settings.UPLOAD_DIR, f"export_{timestamp}.xlsx"
            )

        # 确保输出目录存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "发票列表"

        # 定义表头
        headers = [
            "序号", "发票号码", "发票代码", "发票类型", "开票日期",
            "金额", "税额", "价税合计", "销方名称", "销方税号",
            "购方名称", "购方税号", "校验码", "分类", "状态",
        ]

        # 设置表头样式
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row_idx, invoice in enumerate(invoices, 2):
            row_data = [
                row_idx - 1,
                invoice.invoice_number or "",
                invoice.invoice_code or "",
                invoice.invoice_type or "",
                invoice.invoice_date or "",
                invoice.amount or 0,
                invoice.tax_amount or 0,
                invoice.total_amount or 0,
                invoice.seller_name or "",
                invoice.seller_tax_number or "",
                invoice.buyer_name or "",
                invoice.buyer_tax_number or "",
                invoice.check_code or "",
                invoice.category or "",
                invoice.status or "",
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # 调整列宽
        column_widths = [6, 15, 15, 12, 12, 12, 10, 12, 25, 22, 25, 22, 20, 8, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)].width = width

        # 保存文件
        wb.save(output_path)
        logger.info(f"Excel导出完成: {output_path}，共 {len(invoices)} 条记录")

        return output_path

    async def export_to_pdf(
        self,
        db: AsyncSession,
        invoices: List[Invoice],
        output_path: Optional[str] = None,
    ) -> str:
        """
        将发票数据导出为PDF文件

        Args:
            db: 数据库会话
            invoices: 发票列表
            output_path: 输出文件路径，为None时自动生成

        Returns:
            str: 导出文件的路径
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            raise ImportError("reportlab未安装，请执行: pip install reportlab")

        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(
                settings.UPLOAD_DIR, f"export_{timestamp}.pdf"
            )

        # 确保输出目录存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # 创建PDF文档
        doc = SimpleDocTemplate(
            output_path,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30,
        )

        # 准备样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ChineseTitle",
            parent=styles["Title"],
            fontSize=16,
            spaceAfter=20,
        )

        # 尝试注册中文字体
        try:
            # 尝试常见的中文字体路径
            font_paths = [
                "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
                "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
                "/System/Library/Fonts/PingFang.ttc",
                "C:/Windows/Fonts/simhei.ttf",
                "C:/Windows/Fonts/msyh.ttc",
            ]
            font_registered = False
            for fp in font_paths:
                if os.path.exists(fp):
                    pdfmetrics.registerFont(TTFont("ChineseFont", fp))
                    font_registered = True
                    break

            if not font_registered:
                logger.warning("未找到中文字体文件，PDF中的中文可能无法正常显示")
                chinese_font = "Helvetica"
            else:
                chinese_font = "ChineseFont"
        except Exception as e:
            logger.warning(f"注册中文字体失败: {e}")
            chinese_font = "Helvetica"

        elements = []

        # 添加标题
        title = Paragraph(
            f"Invoice Export Report - {datetime.now().strftime('%Y-%m-%d')}",
            title_style,
        )
        elements.append(title)
        elements.append(Spacer(1, 10))

        # 准备表格数据
        table_headers = [
            "No.", "Number", "Code", "Type", "Date",
            "Amount", "Tax", "Total", "Seller", "Buyer",
            "Category", "Status",
        ]

        table_data = [table_headers]
        for idx, inv in enumerate(invoices, 1):
            row = [
                str(idx),
                inv.invoice_number or "",
                inv.invoice_code or "",
                inv.invoice_type or "",
                inv.invoice_date or "",
                f"{inv.amount or 0:.2f}",
                f"{inv.tax_amount or 0:.2f}",
                f"{inv.total_amount or 0:.2f}",
                (inv.seller_name or "")[:20],
                (inv.buyer_name or "")[:20],
                inv.category or "",
                inv.status or "",
            ]
            table_data.append(row)

        # 创建表格
        col_widths = [30, 80, 80, 60, 70, 60, 50, 60, 120, 120, 50, 50]
        table = Table(table_data, colWidths=col_widths)

        # 设置表格样式
        table.setStyle(TableStyle([
            # 表头样式
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), chinese_font),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # 数据行样式
            ("FONTNAME", (0, 1), (-1, -1), chinese_font),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (0, 1), (-1, -1), "CENTER"),
            # 网格线
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            # 斑马纹
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))

        elements.append(table)

        # 生成PDF
        doc.build(elements)
        logger.info(f"PDF导出完成: {output_path}，共 {len(invoices)} 条记录")

        return output_path

    async def get_invoices_for_export(
        self,
        db: AsyncSession,
        invoice_ids: Optional[List[int]] = None,
        invoice_type: Optional[str] = None,
        category: Optional[str] = None,
        status: Optional[str] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
    ) -> List[Invoice]:
        """
        获取待导出的发票列表

        Args:
            db: 数据库会话
            invoice_ids: 指定导出的发票ID列表
            invoice_type: 按类型筛选
            category: 按分类筛选
            status: 按状态筛选
            start_date: 起始日期
            end_date: 截止日期

        Returns:
            List[Invoice]: 发票列表
        """
        query = select(Invoice)

        if invoice_ids:
            query = query.where(Invoice.id.in_(invoice_ids))
        if invoice_type:
            query = query.where(Invoice.invoice_type == invoice_type)
        if category:
            query = query.where(Invoice.category == category)
        if status:
            query = query.where(Invoice.status == status)
        if start_date:
            query = query.where(Invoice.invoice_date >= start_date)
        if end_date:
            query = query.where(Invoice.invoice_date <= end_date)

        query = query.order_by(Invoice.created_at.desc())

        result = await db.execute(query)
        return list(result.scalars().all())


# 全局导出服务实例
export_service = ExportService()
